import sys as sys
from typing import Any
import openpyxl
import pythoncom
import win32com.client as win32
from flask import (Flask, Response, app, send_file, redirect, render_template,
                   request, url_for)
from openpyxl import load_workbook
from flask_debugtoolbar import DebugToolbarExtension


app = Flask(__name__)

# app.config['SECRET_KEY'] = 'Ameo1988!'
# toolbar = DebugToolbarExtension(app)
items = {}


if __name__ == '__main__':
    app.run(debug=True)


@app.route('/count', methods=['GET', 'POST'])
def count():
    if request.method == 'POST':
        item = request.form['item']
        if item not in items:
            items[item] = 0
    return render_template('count.html', items=items)


@app.route('/increment', methods=['POST'])
def increment():
    item = request.form['item']
    items[item] += 1
    return str(items[item])



@app.route('/', methods=['GET', 'POST'], )  # type: ignore
def form() -> Any:
    # If User hits the Generate button...
    # If Generate button is not pressed, just show index.html
    if request.method == 'POST':
        # Retrieve data from the userinput
        # Assisn user data to a variable
        standard = (request.form['standard'])
        decora = (request.form['decora'])
        gfci = (request.form['gfci'])
        cutin = (request.form['cutin'])
        surface = (request.form['surface'])
        duplex_controlled = (request.form['1switch'])
        
        quad_standard = (request.form['quad_standard'])
        quad_decora = (request.form['quad_decora'])
        quad_gfci = (request.form['quad_gfci'])
        quad_cutin = (request.form['quad_cutin'])
        quad_surface = (request.form['quad_surface'])
        quad_controlled = (request.form['quad_controlled'])

        ff3 = (request.form['3-wire'])
        ff4 = (request.form['4-wire'])

        rough_in_data = (request.form['rough_in_data'])
        cutin_data = (request.form['cutin_data'])


        lv_switch = (request.form['lv_switch'])
        hv_switch = (request.form['hv_switch'])
        hv_dimming = (request.form['hv_dimming'])

        wh_120 = (request.form['wh_120'])
        wh_277 = (request.form['wh_277'])
        wh_277v = (request.form['wh_277v'])

        # Open Excel file
        # Push user iput variable to excel file cell
        # Close Excel file
        wb1 = load_workbook("templates/works.xlsx")
        sheet1 = wb1.active
        sheet1['C5'] = standard  # type: ignore
        sheet1['C6'] = decora  # type: ignore
        sheet1['C7'] = gfci  # type: ignore
        sheet1['C9'] = cutin  # type: ignore
        sheet1['C10'] = surface  # type: ignore
        sheet1['C48'] = duplex_controlled  # type: ignore

        sheet1['C20'] = quad_standard  # type: ignore
        sheet1['C21'] = quad_decora  # type: ignore
        sheet1['C22'] = quad_gfci  # type: ignore
        sheet1['C24'] = quad_cutin  # type: ignore
        sheet1['C51'] = quad_controlled # type: ignore
        sheet1['C23'] = quad_surface  # type: ignore
     

        sheet1['C34'] = ff3  # type: ignore
        sheet1['C35'] = ff4 # type: ignore

        sheet1['C39'] = rough_in_data  # type: ignore   
        sheet1['C40'] = cutin_data  # type: ignore                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                              ccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccc    c
    
        sheet1['C54'] = lv_switch  # type: ignore
        sheet1['C55'] = hv_switch  # type: ignore
        sheet1['C56'] = hv_dimming  # type: ignore

        sheet1['C60'] = wh_120 # type: ignore
        sheet1['C61'] = wh_277 # type: ignore
        sheet1['C61'] = wh_277v # type: ignore


        wb1.save('templates/workstemp.xlsx')
        wb1.close()

        # Redirect to result page
        return redirect(url_for('result'))  # type: ignore
    # show html page
    return render_template('index.html')  # type: ignore


@app.route('/download') # type: ignore
def download() -> Response:
    filename = 'templates/new.xlsx'
    return send_file(filename, as_attachment=True)                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                          
    

@app.route('/c', methods=['GET', 'POST'])
def cool_form():
    if request.method == 'POST':
        # do stuff when the form is submitted

        # redirect to end the POST handling
        # the redirect can be to the same route or somewhere else
        return redirect(url_for('index'))

    # show the form, it wasn't submitted
    return render_template('cool_form.html')

@app.route('/todo')
def todo() -> str:   # show the form, it wasn't submitted
    return render_template('todo.html')
# @app.route('/export')  # type: ignore
# def export() -> str:
#     lists = load_workbook("templates/list.xlsx")
#     list = lists.active
#     # line1 = list.cell(row=11, column=2).value  # type: ignore
#     list['B11'].value = result()[1]  # type: ignore
#     message = "this is a test"
#     list['B12'] = message  # type: ignore
#     print(message)

    # list['B20'] = result()[10]
    # list['B21'] = result()[11]
    # list['B22'] = result()[12]
    # list['B23'] = result()[13]
    # list['B24'] = result()[14]
    # list['B25'] = result()[15]
    # list['B26'] = result()[16]
    # list['B27'] = result()[17]

    # lists.save('templates/listtemp.xlsx')
    # lists.close()
    # file = 'templates/listtemp.xlsx'
    # return send_file(file, as_attachment=True)  # type: ignore


# def load_data():
#     # Load the workbook
#     excel1 = win32.Dispatch("Excel.Application", pythoncom.CoInitialize())
#     wb2 = excel1.Workbooks.Open(r'D:\final\templates\workstemp.xlsx')
#     wb2.Save()
#     wb2.Close()
#     excel1.Quit()
#     wb = load_workbook("templates/workstemp.xlsx", data_only=True)
#     # Select the worksheet and range of cells
#     ws = wb.active
#     range_str = 'H381:I455'

#     # Extract the range of cells as a list of rows
#     rows = ws[range_str]  # type: ignore
#     filtered_rows = []
#     for row in rows:
#         if not all(cell.value is None or cell.value == 0 for cell in row):
#             filtered_rows.append(row)
#     # Render the data and formatting in an HTML table using Jinja2
#     return filtered_rows  # type: ignore


# @app.route('/table')
# def display_range():
#     # Load the data from the Excel file
#     # rows = load_data()
#     # filtered_rows = []
#     # for i, row in enumerate(rows):
#     #   for j, cell in enumerate(row):
#     #     cell_var = f'cell_{i}_{j}'  # Create a new variable name for each cell
#     #     if int(rows[i][1].value) != "0":
#     #         filtered_rows.append(rows[i][0].value)
#     #         filtered_rows.append(rows[i][1].value)
#     #         print(filtered_rows)
#     #     else
#     #  return render_template('table.html', rows=filtered_rows)
#     # Load the Excel file
#     wb = openpyxl.load_workbook('templates/workstemp.xlsx')

#     # Select the first worksheet
#     ws = wb.active

#     # Get all rows in the worksheet
#     rows = list(ws.rows)  # type: ignore

#     # Initialize an empty list to store filtered rows
#     filtered_rows = []

#     # Loop through each cell in the worksheet
#     for i, row in enumerate(rows):
#         for j, cell in enumerate(row):
#             # Create a new variable name for each cell
#             cell_var = f'cell_{i}_{j}'
#             # Assign the cell value to a variable with a dynamic name
#             globals()[cell_var] = cell.value
#             # Check if the value of the second column in the row is not zero
#             if j == 1 and cell.value != 0:
#                 # Append the values of the first and second columns to the filtered rows list
#                 filtered_rows.append(rows[i][0].value)
#                 filtered_rows.append(cell.value)

#     # Render the template with the filtered rows list
#     return render_template('table.html', filtered_rows=filtered_rows)


@app.route('/result')
def result():
    # Open the Excel file to trigger the calculation of formulas
    excel_app = win32.Dispatch("Excel.Application", pythoncom.CoInitialize())
    wb = excel_app.Workbooks.Open(r'D:/Production/templates/workstemp.xlsx')
    wb.Save()
    wb.Close()
    excel_app.Quit()

    # Load the Excel file and read the calculated values
    wb = load_workbook("templates/workstemp.xlsx", data_only=True)
    new = load_workbook("templates/materiallist.xlsx")
    nsheet = new.active
    sheet = wb.active
    row_index = 11
    cell_values = []
    title_values = []
    material = []
    for row in range(381, 456):
        cell_value = sheet.cell(row=row, column=9).value # type: ignore
        title = sheet.cell(row=row, column=8).value # type: ignore
        if cell_value != 0:
        # Append the non-zero cell value to the list
            cell_values.append(cell_value)
            title_values.append(title)
            material.append((title, cell_value)) # type: ignore
            nsheet.cell(row=row_index, column=3, value=title)  # type: ignore # Write title to column 1 of the current row
            nsheet.cell(row=row_index, column=2, value=cell_value)  # type: ignore # Write cell_value to column 2 of the current row
            row_index += 1  # Move down to the next row

    
    rows = []
    starting_row = 11 #Row index to start appending values
    for row in nsheet.iter_rows(min_row=starting_row, values_only=True): # type: ignore
        rows.append(row)
        print(row)
    new.save("templates/new.xlsx")
    new.close()
    
    wb.save("templates/workstemp.xlsx")
    wb.close()

    # Pass the cell values to the template
    return render_template('result.html', rows=rows)
    # return   send_file(filename, as_attachment=True, ) # type: ignore


@app.route('/manifest.json')  # type: ignore
def manifest() -> Response:
    return app.send_static_file('manifest.json')

# Serve service worker file


@app.route('/sw.js')
def service_worker() -> Response:
    return app.send_static_file('sw.js')

# Cache static assets


@app.after_request
def add_header(response) -> Response:
    response.headers['Cache-Control'] = 'static, max-age=31536000'
    return response


if __name__ == '__main__':
    app.run(debug=True)
