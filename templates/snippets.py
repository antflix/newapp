
##################read data from excel file and print it on site as table#########################
from flask import app, render_template
from openpyxl import load_workbook
import win32com


def load_data():
    # Load the workbook
    excel1 = win32com.Dispatch("Excel.Application", pythoncom.CoInitialize()) # type: ignore
    wb2 = excel1.Workbooks.Open(r'D:\final\templates\workstemp.xlsx')
    wb2.Save()
    wb2.Close()
    excel1.Quit()
    wb = load_workbook("templates/workstemp.xlsx", data_only=True)
    # Select the worksheet and range of cells
    ws = wb.active
    range_str = 'H381:I455'
    
    # Extract the range of cells as a list of rows
    rows = ws[range_str] # type: ignore
    filtered_rows = []
    for row
    in rows:
        if not all(cell.value is None or cell.value == 0 for cell in row):
            filtered_rows.append(row)
    # Render the data and formatting in an HTML table using Jinja2
    return filtered_rows # type: ignore

@app.route('/table') # type: ignore
def display_range():
    # Load the data from the Excel file
    rows = load_data()
    # Render the data and formatting in an HTML table using Jinja2
    return render_template('table.html', rows=rows)